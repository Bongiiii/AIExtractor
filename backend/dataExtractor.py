import os
import io
import json
import base64
import time
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from PIL import Image
import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load API key from .env file
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY not found in .env file. Please add OPENAI_API_KEY=your_key_here to your .env file")


class EnhancedPDFExtractor:
    def __init__(self, api_key: str):
        """
        Initialize the Enhanced PDF Table Extractor
        
        Args:
            api_key (str): OpenAI API key
        """
        self.client = OpenAI(api_key=api_key)
        self.output_dir = "extracted_tables"
        os.makedirs(self.output_dir, exist_ok=True)
        logger.info(f"Output directory: {os.path.abspath(self.output_dir)}")
    
    def pdf_to_images(self, pdf_path: str, dpi: int = 200) -> List[Image.Image]:
        """
        Convert PDF pages to images with optimized DPI for text recognition
        
        Args:
            pdf_path (str): Path to PDF file
            dpi (int): Resolution for image conversion
            
        Returns:
            List[Image.Image]: List of PIL Images
        """
        try:
            doc = fitz.open(pdf_path)
            images = []
            
            for page_num in range(len(doc)):
                try:
                    page = doc.load_page(page_num)
                    mat = fitz.Matrix(dpi/72, dpi/72)  # Scale factor for DPI
                    pix = page.get_pixmap(matrix=mat)
                    img_data = pix.tobytes("png")
                    img = Image.open(io.BytesIO(img_data))
                    images.append(img)
                except Exception as e:
                    logger.error(f"Error processing page {page_num + 1}: {e}")
                    continue
            
            doc.close()
            logger.info(f"Successfully converted {len(images)} pages to images")
            return images
            
        except Exception as e:
            logger.error(f"Error opening PDF {pdf_path}: {e}")
            raise
    
    def encode_image(self, image: Image.Image) -> str:
        """
        Encode PIL Image to base64 string
        
        Args:
            image (Image.Image): PIL Image
            
        Returns:
            str: Base64 encoded image
        """
        buffer = io.BytesIO()
        image.save(buffer, format='PNG')
        image_bytes = buffer.getvalue()
        return base64.b64encode(image_bytes).decode('utf-8')
    
    def extract_dense_table_data(self, image: Image.Image, columns: List[str], 
                               extra_instructions: str = "", page_num: int = 0) -> List[Dict[str, Any]]:
        """
        Enhanced extraction specifically for dense, multi-column scientific documents
        
        Args:
            image (Image.Image): PIL Image of the document page
            columns (List[str]): List of column names to extract
            extra_instructions (str): Additional instructions for extraction
            page_num (int): Page number for reference
            
        Returns:
            List[Dict]: Extracted data rows
        """
        try:
            base64_image = self.encode_image(image)
            
            columns_str = '", "'.join(columns)
            
            # Enhanced prompt specifically for dense scientific documents
            prompt = f"""
            CRITICAL: This appears to be a dense, multi-column scientific document with species/taxonomic data. 
            Your task is to extract ALL tabular data that matches these columns: "{columns_str}"

            DOCUMENT ANALYSIS INSTRUCTIONS:
            1. This document likely contains species/taxonomic data in a structured format
            2. Look for patterns like: Scientific names (Latin binomials), Common names, Location data, Status codes
            3. Data may be organized in multiple columns across the page
            4. Some rows may span multiple lines due to long scientific names or locations
            5. Ignore headers like "PROPOSED RULE MAKING", page numbers, and section titles
            6. Focus on the actual data entries, not the formatting elements

            EXTRACTION STRATEGY:
            - Scan the ENTIRE page systematically from top to bottom
            - Look for repeating patterns of data that match the requested columns
            - Scientific names often follow the pattern: "Genus species" (italicized or not)
            - Location data might include counties, states, countries
            - Status information might be coded (like "Ex", "Extinct", "Threatened", etc.)
            - Extract from ALL visible tabular content

            SPECIFIC COLUMN MATCHING:
            {self._generate_enhanced_column_definitions(columns)}

            DATA QUALITY RULES:
            1. Extract EVERY row that contains relevant data - prioritize completeness
            2. If a field spans multiple lines, combine the text appropriately
            3. Preserve scientific naming conventions exactly
            4. Keep location information as detailed as possible
            5. If status codes are abbreviated, keep them as-is
            6. For empty/missing fields, use "N/A"
            7. Remove any obvious formatting artifacts or page headers/footers

            CRITICAL OUTPUT FORMAT:
            Return a JSON object with this structure:
            {{
                "extracted_data": [
                    {{{", ".join([f'"{col}": "extracted_value"' for col in columns])}}},
                    {{{", ".join([f'"{col}": "extracted_value"' for col in columns])}}},
                    ... (continue for ALL data rows found)
                ],
                "total_rows": actual_number_of_rows_extracted,
                "extraction_notes": "observations about data density, formatting challenges, etc.",
                "confidence_level": "high/medium/low based on text clarity"
            }}

            ADDITIONAL CONTEXT:
            {extra_instructions if extra_instructions else "No additional context provided."}

            IMPORTANT: Dense scientific documents often contain many rows per page. 
            Make sure you're not missing large sections of data. Be thorough and systematic.
            If you find obvious tabular structures, extract ALL rows from them.
            """
            
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=4000,
                temperature=0.05
            )
            
            # Parse the JSON response
            response_text = response.choices[0].message.content
            logger.info(f"Raw response from OpenAI (Page {page_num + 1}): {response_text[:200]}...")
            
            # Clean up the response to extract JSON
            if "```json" in response_text:
                json_start = response_text.find("```json") + 7
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            elif "```" in response_text:
                json_start = response_text.find("```") + 3
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            
            try:
                result = json.loads(response_text)
                extracted_data = result.get("extracted_data", [])
                
                logger.info(f"Successfully extracted {len(extracted_data)} rows from page {page_num + 1}")
                if result.get("extraction_notes"):
                    logger.info(f"Notes: {result.get('extraction_notes')}")
                if result.get("confidence_level"):
                    logger.info(f"Confidence: {result.get('confidence_level')}")
                
                return extracted_data
            
            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error on page {page_num + 1}: {e}")
                logger.info("Attempting to parse partial response...")
                return self._parse_partial_response(response_text, columns)
        
        except Exception as e:
            logger.error(f"Error processing page {page_num + 1}: {e}")
            return []
    
    def _parse_partial_response(self, response_text: str, columns: List[str]) -> List[Dict[str, Any]]:
        """
        Attempt to parse partial or malformed JSON responses
        """
        try:
            import re
            
            # Look for array patterns
            array_pattern = r'\[.*?\]'
            arrays = re.findall(array_pattern, response_text, re.DOTALL)
            
            for array_str in arrays:
                try:
                    data = json.loads(array_str)
                    if isinstance(data, list) and data and isinstance(data[0], dict):
                        logger.info(f"Recovered {len(data)} rows from partial response")
                        return data
                except:
                    continue
            
            # If no arrays found, try to extract individual objects
            object_pattern = r'\{[^{}]*\}'
            objects = re.findall(object_pattern, response_text)
            
            recovered_data = []
            for obj_str in objects:
                try:
                    obj = json.loads(obj_str)
                    if isinstance(obj, dict) and any(col in obj for col in columns):
                        recovered_data.append(obj)
                except:
                    continue
            
            if recovered_data:
                logger.info(f"Recovered {len(recovered_data)} rows from individual objects")
                return recovered_data
                
        except Exception as e:
            logger.error(f"Failed to parse partial response: {e}")
        
        return []
    
    def _generate_enhanced_column_definitions(self, columns: List[str]) -> str:
        """Generate enhanced definitions for scientific document columns"""
        definitions = []
        for col in columns:
            col_lower = col.lower()
            if any(word in col_lower for word in ['species', 'scientific', 'name', 'binomial', 'taxa']):
                definitions.append(f"- {col}: Scientific binomial names (e.g., 'Quercus alba', 'Homo sapiens') - look for Latin genus + species")
            elif any(word in col_lower for word in ['common', 'vernacular', 'english']):
                definitions.append(f"- {col}: Common English names (e.g., 'White Oak', 'American Robin')")
            elif any(word in col_lower for word in ['location', 'locality', 'place', 'county', 'state', 'where', 'found', 'range']):
                definitions.append(f"- {col}: Geographic information (counties, states, countries, specific localities)")
            elif any(word in col_lower for word in ['date', 'year', 'time', 'collected', 'observed', 'when']):
                definitions.append(f"- {col}: Temporal information (dates, years, time periods)")
            elif any(word in col_lower for word in ['status', 'condition', 'conservation', 'threat', 'endangered', 'extinct']):
                definitions.append(f"- {col}: Conservation/threat status (e.g., 'Extinct', 'Endangered', 'Ex', 'En', status codes)")
            elif any(word in col_lower for word in ['family', 'group', 'category', 'class', 'order']):
                definitions.append(f"- {col}: Taxonomic classification (Family, Order, Class names)")
            else:
                definitions.append(f"- {col}: Data that logically corresponds to this column name")
        
        return "\n".join(definitions)
    
    def process_pdf_enhanced(self, pdf_path: str, columns: List[str], 
                           extra_instructions: str = "", extract_multiple_rows: bool = True,
                           resume_from_page: int = 0, sample_pages: int = None) -> str:
        """
        Enhanced PDF processing with better error handling and validation
        
        Args:
            pdf_path (str): Path to the PDF file
            columns (List[str]): List of column names to extract
            extra_instructions (str): Additional instructions for extraction
            extract_multiple_rows (bool): Whether to extract multiple rows per document
            resume_from_page (int): Page number to resume from (0-indexed)
            sample_pages (int): If provided, only process this many pages (for testing)
            
        Returns:
            str: Path to the created Excel file
        """
        pdf_name = Path(pdf_path).stem
        columns_suffix = "_".join(col.replace(" ", "")[:3] for col in columns[:3])
        output_excel_path = os.path.join(self.output_dir, f"{pdf_name}_{columns_suffix}_enhanced.xlsx")
        
        logger.info(f"Processing PDF: {pdf_path}")
        logger.info(f"Extracting columns: {', '.join(columns)}")
        
        try:
            # Load previous progress if resuming
            all_data, prev_columns = self.load_progress(pdf_name)
            if all_data and prev_columns == columns:
                logger.info(f"Resuming from previous progress ({len(all_data)} rows already extracted)")
            else:
                all_data = []
                resume_from_page = 0
            
            # Convert PDF to images
            logger.info("Converting PDF to images...")
            images = self.pdf_to_images(pdf_path, dpi=200)
            total_pages = len(images)
            
            if sample_pages:
                total_pages = min(sample_pages, total_pages)
                images = images[:sample_pages]
                logger.info(f"Processing sample of {total_pages} pages for testing")
            
            logger.info(f"Successfully converted PDF to {total_pages} images")
            
            if resume_from_page > 0:
                logger.info(f"Resuming from page {resume_from_page + 1}")
            
            # Process each page with enhanced extraction
            pages_with_data = 0
            pages_without_data = 0
            total_rows_extracted = len(all_data)
            
            for page_num in range(resume_from_page, total_pages):
                progress_percent = ((page_num + 1) / total_pages) * 100
                logger.info(f"Processing page {page_num + 1}/{total_pages} ({progress_percent:.1f}%)")
                
                try:
                    image = images[page_num]
                    
                    # Use enhanced extraction method
                    page_data = self.extract_dense_table_data(
                        image, columns, extra_instructions, page_num
                    )
                    
                    if page_data:
                        # Add page number to each row
                        for row in page_data:
                            row['_page_number'] = page_num + 1
                        
                        all_data.extend(page_data)
                        pages_with_data += 1
                        total_rows_extracted += len(page_data)
                        logger.info(f"Extracted {len(page_data)} rows from page {page_num + 1}")
                        logger.info(f"Running total: {total_rows_extracted} rows")
                    else:
                        pages_without_data += 1
                        logger.info(f"No data found on page {page_num + 1}")
                    
                    # Save progress more frequently for large extractions
                    if (page_num + 1) % 5 == 0:
                        self.save_progress(all_data, pdf_name, columns)
                        logger.info(f"Progress saved ({len(all_data)} total rows)")
                    
                    # Adaptive rate limiting
                    if page_num < total_pages - 1:
                        if len(page_data) > 50:
                            time.sleep(2)
                        elif len(page_data) > 0:
                            time.sleep(1)
                        else:
                            time.sleep(0.5)
                        
                except Exception as e:
                    logger.error(f"Error processing page {page_num + 1}: {e}")
                    continue
            
            # Enhanced final statistics
            logger.info("EXTRACTION SUMMARY:")
            logger.info(f"Total pages processed: {total_pages}")
            logger.info(f"Pages with data: {pages_with_data}")
            logger.info(f"Pages without data: {pages_without_data}")
            if total_pages > 0:
                logger.info(f"Success rate: {(pages_with_data/total_pages)*100:.1f}%")
            logger.info(f"Total rows extracted: {len(all_data)}")
            
            if pages_with_data > 0:
                avg_rows_per_page = len(all_data) / pages_with_data
                logger.info(f"Average rows per successful page: {avg_rows_per_page:.1f}")
            
            # Remove page number column if user didn't request it
            if '_page_number' not in columns:
                for row in all_data:
                    row.pop('_page_number', None)
            
            if all_data:
                logger.info(f"Creating Excel file with {len(all_data)} total rows...")
                self.create_excel_file(all_data, columns, output_excel_path, extract_multiple_rows)
                
                # Clean up progress file on success
                self.clean_progress(pdf_name)
                
                logger.info(f"Successfully extracted data from {len(all_data)} rows!")
                
                # Data quality assessment
                self._assess_data_quality(all_data, columns)
                
            else:
                logger.warning("No data found in the entire document")
                # Create empty Excel file
                pd.DataFrame(columns=columns).to_excel(output_excel_path, index=False)
            
            return output_excel_path
            
        except Exception as e:
            logger.error(f"Error processing PDF {pdf_path}: {e}")
            logger.info("Progress has been saved. You can resume processing later.")
            raise
    
    def _assess_data_quality(self, data: List[Dict], columns: List[str]):
        """Assess and report data quality metrics"""
        if not data:
            return
        
        logger.info("DATA QUALITY ASSESSMENT:")
        
        total_rows = len(data)
        for col in columns:
            non_empty = sum(1 for row in data if row.get(col) and str(row[col]).strip() not in ['', 'N/A', 'n/a', 'nan'])
            completeness = (non_empty / total_rows) * 100
            logger.info(f"{col}: {completeness:.1f}% complete ({non_empty}/{total_rows} rows)")
        
        # Check for potential data patterns
        if data:
            sample_row = data[0]
            logger.info("SAMPLE ROW:")
            for col in columns:
                value = sample_row.get(col, 'N/A')
                logger.info(f"{col}: {value}")
    
    def create_excel_file(self, all_data: List[Dict[str, Any]], columns: List[str], 
                         output_path: str, extract_multiple_rows: bool = True):
        """Enhanced Excel creation with better formatting and data validation"""
        if not all_data:
            logger.warning("No data to write to Excel file")
            pd.DataFrame(columns=columns).to_excel(output_path, index=False)
            return
        
        # Create DataFrame with better handling of missing columns
        df = pd.DataFrame(all_data)
        
        # Ensure all specified columns exist
        for col in columns:
            if col not in df.columns:
                df[col] = "N/A"
        
        # Reorder columns and clean data
        df = df[columns].copy()
        
        # Clean up data
        for col in columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace(['nan', 'None', 'null'], 'N/A')
        
        # Create Excel file with enhanced formatting
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Extracted_Data', index=False)
                
                # Enhanced formatting
                worksheet = writer.sheets['Extracted_Data']
                
                # Style headers
                for col in range(1, len(columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_length = len(str(cell.value)) if cell.value else 0
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = min(max(max_length + 3, 15), 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            # Fallback to basic Excel creation
            df.to_excel(output_path, index=False)
        
        logger.info(f"Excel file created: {output_path}")
        logger.info(f"Total rows: {len(df)}")
    
    def save_progress(self, data: List[Dict], pdf_name: str, columns: List[str]):
        """Enhanced progress saving with metadata"""
        try:
            progress_file = os.path.join(self.output_dir, f"{pdf_name}_progress.json")
            progress_data = {
                'data': data,
                'columns': columns,
                'timestamp': time.time(),
                'total_rows': len(data)
            }
            with open(progress_file, 'w') as f:
                json.dump(progress_data, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving progress: {e}")
    
    def load_progress(self, pdf_name: str) -> tuple:
        """Enhanced progress loading with validation"""
        progress_file = os.path.join(self.output_dir, f"{pdf_name}_progress.json")
        if os.path.exists(progress_file):
            try:
                with open(progress_file, 'r') as f:
                    progress = json.load(f)
                    return progress.get('data', []), progress.get('columns', [])
            except Exception as e:
                logger.error(f"Error loading progress: {e}")
        return [], []
    
    def clean_progress(self, pdf_name: str):
        """Clean up progress file"""
        progress_file = os.path.join(self.output_dir, f"{pdf_name}_progress.json")
        if os.path.exists(progress_file):
            try:
                os.remove(progress_file)
            except Exception as e:
                logger.error(f"Error cleaning progress file: {e}")


def main_enhanced():
    """
    Enhanced main function with better user experience
    """
    INPUT_DIRECTORY = "input_pdfs"
    
    print("\n" + "="*70)
    print(" ENHANCED PDF TABLE EXTRACTOR")
    print("   Optimized for Dense Scientific Documents")
    print("="*70)
    
    # Check API key
    if not OPENAI_API_KEY:
        print(" Error: OpenAI API key not found!")
        print("Please create a .env file with: OPENAI_API_KEY=your_key_here")
        return
    
    # Create input directory
    os.makedirs(INPUT_DIRECTORY, exist_ok=True)
    
    # Check for PDF files
    pdf_files = list(Path(INPUT_DIRECTORY).glob("*.pdf"))
    
    if not pdf_files:
        print(f" No PDF files found in {INPUT_DIRECTORY}")
        print(f" Please place your PDFs in: {os.path.abspath(INPUT_DIRECTORY)}")
        return
    
    print(f" Found {len(pdf_files)} PDF file(s)")
    
    # Interactive setup with better defaults for scientific documents
    print("\n COLUMN SETUP:")
    print("Enter column names (common for scientific docs: Species, Common Name, Location, Status)")
    
    columns = []
    suggested_columns = ["Species", "Common Name", "Location", "Status", "Date"]
    
    print(f"\n Suggested columns: {', '.join(suggested_columns)}")
    use_suggested = input("Use suggested columns? (Y/n): ").strip().lower()
    
    if use_suggested != 'n':
        columns = suggested_columns
    else:
        print("\nEnter custom column names (press Enter on empty line to finish):")
        while True:
            col = input(f"Column {len(columns) + 1}: ").strip()
            if not col:
                break
            columns.append(col)
    
    if not columns:
        columns = suggested_columns
    
    print(f"\n Final columns: {', '.join(columns)}")
    
    # Additional instructions
    print("\n ADDITIONAL INSTRUCTIONS (optional):")
    extra_instructions = input("Any specific extraction notes: ").strip()
    
    # Test mode option
    print("\n TEST MODE:")
    test_mode = input("Run in test mode (process only first 3 pages)? (y/N): ").strip().lower()
    sample_pages = 3 if test_mode == 'y' else None
    
    # Initialize enhanced extractor
    try:
        extractor = EnhancedPDFExtractor(OPENAI_API_KEY)
        
        for pdf_file in pdf_files:
            print(f"\n{'='*70}")
            print(f" PROCESSING: {pdf_file.name}")
            print('='*70)
            
            try:
                result = extractor.process_pdf_enhanced(
                    str(pdf_file), 
                    columns, 
                    extra_instructions,
                    extract_multiple_rows=True,
                    sample_pages=sample_pages
                )
                
                if result:
                    print(f" Success: {pdf_file.name}")
                    print(f" Output: {result}")
                else:
                    print(f" Failed: {pdf_file.name}")
                    
            except KeyboardInterrupt:
                print(f"\n Processing interrupted by user")
                break
            except Exception as e:
                print(f" Error: {e}")
        
        print(f"\n Processing complete!")
        print(f" Check '{extractor.output_dir}' for results")
        
    except Exception as e:
        print(f" Error initializing extractor: {e}")


if __name__ == "__main__":
    main_enhanced()